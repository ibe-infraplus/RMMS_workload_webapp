import openpyxl

def add_bridge_m_column():
    file_path = 'data/group_data_final.xlsx'
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    # Target sheet
    sheet_name = 'group_data_final'
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        return
        
    ws = wb[sheet_name]
    
    # We want to insert 'bridge_m' at column AN, which is the 40th column.
    # In openpyxl, columns are 1-indexed. A=1, Z=26, AA=27, AN=40.
    col_idx = 40
    
    # Insert a new column at index 40
    ws.insert_cols(col_idx)
    
    # Set header
    ws.cell(row=1, column=col_idx, value='bridge_m')
    
    # Fill remaining rows with 0
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=col_idx, value=0)
        
    print(f"Added 'bridge_m' column at AN (index 40) with 0s for {max_row-1} rows.")
    
    wb.save(file_path)
    print("Workbook saved successfully.")

if __name__ == "__main__":
    add_bridge_m_column()
